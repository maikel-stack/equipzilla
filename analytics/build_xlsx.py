#!/usr/bin/env python3
"""Generate equipzilla_events_plan.xlsx from the three CSVs in analytics/."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
SHEETS = [
    ("Plan", "equipzilla_events_plan.csv"),
    ("Parametros", "equipzilla_events_params.csv"),
    ("Backlog", "equipzilla_events_backlog.csv"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Inter", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Inter", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PRIORITY_FILLS = {
    "P0": PatternFill("solid", fgColor="FEE2E2"),
    "P1": PatternFill("solid", fgColor="FEF3C7"),
    "P2": PatternFill("solid", fgColor="DBEAFE"),
}
STATUS_FILLS = {
    "Pending": PatternFill("solid", fgColor="F3F4F6"),
    "Dev": PatternFill("solid", fgColor="FEF3C7"),
    "QA": PatternFill("solid", fgColor="DBEAFE"),
    "Live": PatternFill("solid", fgColor="D1FAE5"),
    "Blocked": PatternFill("solid", fgColor="FEE2E2"),
}

COL_WIDTHS = {
    "Plan": [14, 24, 40, 26, 24, 50, 32, 10, 14, 12, 14, 40],
    "Parametros": [18, 12, 38, 32, 26, 38],
    "Backlog": [10, 38, 60, 36, 60, 12, 10, 18, 12, 16],
}

wb = Workbook()
wb.remove(wb.active)

for sheet_name, csv_name in SHEETS:
    ws = wb.create_sheet(sheet_name)
    rows = list(csv.reader((ROOT / csv_name).open(encoding="utf-8")))
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = WRAP
            cell.border = BORDER
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            else:
                cell.font = BODY_FONT

    headers = rows[0]
    for r_idx in range(2, len(rows) + 1):
        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if header == "Prioridad" and cell.value in PRIORITY_FILLS:
                cell.fill = PRIORITY_FILLS[cell.value]
            elif header == "Estado" and cell.value in STATUS_FILLS:
                cell.fill = STATUS_FILLS[cell.value]

    widths = COL_WIDTHS.get(sheet_name, [20] * len(headers))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 100

    last_col = get_column_letter(len(headers))
    last_row = len(rows)
    table_ref = f"A1:{last_col}{last_row}"
    table = Table(displayName=f"tbl_{sheet_name}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight15",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

readme = wb.create_sheet("Leeme", 0)
readme.column_dimensions["A"].width = 100
readme["A1"] = "Plan de eventos de conversion - Equipzilla"
readme["A1"].font = Font(name="Inter", size=16, bold=True, color="1F2937")
intro = [
    "",
    "Este libro contiene tres pestanas:",
    "  - Plan: catalogo de 28 eventos a implementar (con payload, ubicacion y prioridad).",
    "  - Parametros: glosario de los 39 parametros usados en los eventos (tipo, formato, ejemplo).",
    "  - Backlog: 10 tickets de implementacion (ANL-1 a ANL-10) con criterios de aceptacion.",
    "",
    "Como usarlo:",
    "  1. Sube este .xlsx a Google Drive y abrelo: se convierte en Google Sheets nativo.",
    "  2. Comparte desde el boton 'Compartir' como cualquier hoja.",
    "  3. En las columnas Estado y Prioridad puedes anadir Data Validation (Dropdown) para que sea editable.",
    "",
    "Reglas inviolables:",
    "  - Ningun parametro contiene PII (email, telefono, nombre crudo).",
    "  - Todos los eventos pasan por el helper track() del front, no por gtag directo.",
    "  - generate_lead y sign_up se duplican server-side (ANL-9).",
    "  - Consent Mode v2 con default denied para region ES.",
    "",
    "Fuente: rama claude/equipzilla-conversion-events-oIFdl del repo equipzilla.",
]
for i, line in enumerate(intro, start=2):
    readme.cell(row=i, column=1, value=line).font = Font(name="Inter", size=11)

wb.save(ROOT / "equipzilla_events_plan.xlsx")
print("OK")
